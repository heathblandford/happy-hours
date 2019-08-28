import openpyxl
import geopy
from geopy.geocoders import Nominatim
import time

# set up work sheet globals
wb = openpyxl.load_workbook('barList.xlsx')
sheet = wb.get_sheet_by_name('barList')
geolocator = Nominatim(user_agent="thing_to_getLatandLong")
barListMaxRow = sheet.max_row

for rowNum in sheet.iter_rows(min_row=2, max_row=barListMaxRow, min_col=5, max_col=5):
    for cell in rowNum:
        location = geolocator.geocode(cell.value)
        if location is None:
            print("none", cell.coordinate)
        else:
            currRow = cell.row
            sheet.cell(row=currRow, column=23).value = location.latitude
            sheet.cell(row=currRow, column=24).value = location.longitude
        time.sleep(1.5)

wb.save('barListAppended.xlsx')